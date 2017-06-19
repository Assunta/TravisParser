import pymysql
from openpyxl import load_workbook

from checkTaskGradle import readDbLogin

credenziali = readDbLogin()
connection = pymysql.connect(host=credenziali[0],
                             user=credenziali[1],
                             password=credenziali[2],
                             db='travisdb',
                             charset='utf8mb4',
                             cursorclass=pymysql.cursors.DictCursor)

# try:
#     # with connection.cursor() as cursor:
#     #     sql = "SELECT `*`FROM `category`"
#     #     cursor.execute(sql)
#     #     for r in cursor.fetchall():
#     #         print(r)
#     #
#     # with connection.cursor() as cursor:
#     #     # Read a single record
#     #     sql = "SELECT `Goal`, `Category` FROM `goalmaven` WHERE `Goal`=%s"
#     #     cursor.execute(sql, ('org.apache.maven.plugins:maven-plugin-plugin:helpmojo',))
#     #     result = cursor.fetchone()
#     #     print(result)
#
#     wb = load_workbook('C:\\Users\\Assunta\\Desktop\\TESI\\taskGradle.xlsx')
#     sheets = wb.sheetnames
#     ws = wb[sheets[0]]
#     Myrange = range(1, 12, 1)
#     for i in Myrange:
#         #id = "A" + str(i)
#         cat= "C"+str(i)
#         nome = "B" + str(i)
#         with connection.cursor() as cursor:
#             # Create a new record
#             print (ws[cat].value)
#             sql = "INSERT INTO `regoletaskgradle` (`EspressioneRegolare`,`Categoria`) VALUES ( %s, %s)"
#             cursor.execute(sql, ((ws[nome].value), (ws[cat].value)))
#             # connection is not autocommit by default. So you must commit to save
#             # your changes.
#         connection.commit()
#
#
#
#
# finally:
#     connection.close()